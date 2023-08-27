from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Carga spreadsheet ya existente
wb = load_workbook("C:/Users/mgangui/OneDrive - BVS TELEVISION SRL/Escritorio/cisco.xlsx")

# Crea una spreadsheet activa
ws = wb.active

# Printea A2
print("Equipo: ", ws["A2"].value)
print("GPL price: ", ws["B2"].value)

# Otra manera mas prolija es:

Equipo =  ws["A2"].value
GPL_Price = ws["B2"].value

print(f"{Equipo}: {GPL_Price}") 



#Agarrar una columna entera

columna_b = ws["B"]
columna_c = ws["C"]
for cell in columna_b:
    #print(cell.value)
    break
    

# agarremos un rango
range = ws["A2":"A10"]
print(range)
for i in range:
    print(i) # como es una tupla tenemos que hacer otro loop
    for j in i:
        print(j.value)








