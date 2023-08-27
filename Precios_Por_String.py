import openpyxl
from openpyxl import load_workbook

# Carga spreadsheet ya existente
wb = load_workbook("C:/Users/mgangui/OneDrive - BVS TELEVISION SRL/Escritorio/cisco.xlsx")

# Crea una spreadsheet activa
ws = wb.active


# Itera a través de las celdas en la columna B
for cell in ws['A']:
    if "WS-C2960-48PST-S" in str(cell.value):               #if "escribir el string que ando buscando" 

        # Obtiene la celda correspondiente en la columna C
        celda_columna_c = ws.cell(row=cell.row, column=3)

        # Establece el valor en la celda de la columna C
        celda_columna_c.value = 99

# Guarda los cambios en el archivo
wb.save("C:/Users/mgangui/OneDrive - BVS TELEVISION SRL/Escritorio/cisco.xlsx")
wb = openpyxl.load_workbook("C:/Users/mgangui/OneDrive - BVS TELEVISION SRL/Escritorio/cisco.xlsx")